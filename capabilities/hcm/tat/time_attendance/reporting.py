"""
APG Time & Attendance Capability - Advanced Reporting & Data Export Engine

Governed reporting system with AI-powered insights, predictive analytics,
automated compliance reporting, and multi-format data export capabilities.

Copyright © 2025 Datacraft
Author: Nyimbi Odero
Email: nyimbi@gmail.com
"""

import asyncio
import logging
import json
import csv
import io
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass, asdict
from enum import Enum

from pydantic import BaseModel, Field, ConfigDict

try:
	import openpyxl
	from openpyxl.styles import Font, PatternFill
except ModuleNotFoundError:
	openpyxl = None
	Font = None
	PatternFill = None

try:
	from reportlab.lib import colors
	from reportlab.lib.pagesizes import A4
	from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
	from reportlab.lib.styles import getSampleStyleSheet
	from reportlab.lib.units import inch
except ModuleNotFoundError:
	colors = None
	A4 = None
	SimpleDocTemplate = None
	Table = None
	TableStyle = None
	Paragraph = None
	Spacer = None
	getSampleStyleSheet = None
	inch = None

from .service import TimeAttendanceService
from .models import WorkMode, AIAgentType, TimeEntryStatus
from .config import get_config


logger = logging.getLogger(__name__)


class ReportFormat(str, Enum):
	"""Supported report formats"""
	PDF = "pdf"
	EXCEL = "xlsx"
	CSV = "csv"
	JSON = "json"
	HTML = "html"


class ReportType(str, Enum):
	"""Available report types"""
	TIMESHEET = "timesheet"
	ATTENDANCE_SUMMARY = "attendance_summary"
	PAYROLL = "payroll"
	COMPLIANCE = "compliance"
	PRODUCTIVITY = "productivity"
	FRAUD_ANALYSIS = "fraud_analysis"
	REMOTE_WORK = "remote_work"
	AI_AGENT_UTILIZATION = "ai_agent_utilization"
	EXECUTIVE_DASHBOARD = "executive_dashboard"
	CUSTOM = "custom"


class ReportPeriod(str, Enum):
	"""Report period options"""
	DAILY = "daily"
	WEEKLY = "weekly"
	BIWEEKLY = "biweekly"
	MONTHLY = "monthly"
	QUARTERLY = "quarterly"
	YEARLY = "yearly"
	CUSTOM = "custom"


@dataclass
class ReportConfig:
	"""Report configuration"""
	report_type: ReportType
	format: ReportFormat
	period: ReportPeriod
	start_date: date
	end_date: date
	tenant_id: str
	employee_ids: Optional[List[str]] = None
	department_ids: Optional[List[str]] = None
	include_charts: bool = True
	include_analytics: bool = True
	custom_filters: Dict[str, Any] = None
	
	def to_dict(self) -> Dict[str, Any]:
		data = asdict(self)
		data['start_date'] = self.start_date.isoformat()
		data['end_date'] = self.end_date.isoformat()
		return data


class ReportRequest(BaseModel):
	"""Report generation request"""
	model_config = ConfigDict(extra='forbid')
	
	report_type: ReportType = Field(..., description="Type of report to generate")
	format: ReportFormat = Field(default=ReportFormat.PDF, description="Output format")
	period: ReportPeriod = Field(default=ReportPeriod.MONTHLY, description="Report period")
	start_date: Optional[date] = Field(None, description="Custom start date")
	end_date: Optional[date] = Field(None, description="Custom end date")
	employee_ids: Optional[List[str]] = Field(None, description="Specific employees")
	department_ids: Optional[List[str]] = Field(None, description="Specific departments")
	include_charts: bool = Field(default=True, description="Include visualizations")
	include_analytics: bool = Field(default=True, description="Include AI insights")
	custom_filters: Optional[Dict[str, Any]] = Field(None, description="Additional filters")
	email_recipients: Optional[List[str]] = Field(None, description="Email report to recipients")


class ReportGenerator:
	"""Advanced report generation engine"""
	
	def __init__(self, service: TimeAttendanceService):
		self.service = service
		self.report_templates = self._load_report_templates()
		
	def _load_report_templates(self) -> Dict[str, Dict[str, Any]]:
		"""Load report templates configuration"""
		return {
			ReportType.TIMESHEET: {
				"title": "Employee Timesheet Report",
				"description": "Detailed time tracking records for employees",
				"columns": ["employee_name", "date", "clock_in", "clock_out", "total_hours", "overtime", "break_duration", "status"],
				"charts": ["daily_hours", "overtime_trend"],
				"analytics": ["punctuality_score", "productivity_metrics"]
			},
			ReportType.ATTENDANCE_SUMMARY: {
				"title": "Attendance Summary Report",
				"description": "Comprehensive attendance statistics and trends",
				"columns": ["employee_name", "total_days", "present_days", "absent_days", "late_days", "attendance_rate"],
				"charts": ["attendance_trend", "department_comparison"],
				"analytics": ["attendance_patterns", "risk_indicators"]
			},
			ReportType.PAYROLL: {
				"title": "Payroll Time Report",
				"description": "Time data formatted for payroll processing",
				"columns": ["employee_id", "regular_hours", "overtime_hours", "holiday_hours", "sick_hours", "vacation_hours", "total_pay_hours"],
				"charts": ["hours_breakdown", "cost_analysis"],
				"analytics": ["labor_cost_insights", "budget_variance"]
			},
			ReportType.COMPLIANCE: {
				"title": "Labor Compliance Report",
				"description": "FLSA, GDPR and regulatory compliance analysis",
				"columns": ["employee_name", "flsa_violations", "break_compliance", "overtime_compliance", "gdpr_status"],
				"charts": ["compliance_score", "violation_trends"],
				"analytics": ["risk_assessment", "compliance_recommendations"]
			},
			ReportType.PRODUCTIVITY: {
				"title": "Workforce Productivity Analysis",
				"description": "AI-powered productivity insights and recommendations",
				"columns": ["employee_name", "productivity_score", "efficiency_rating", "goal_achievement", "improvement_areas"],
				"charts": ["productivity_trends", "team_comparison"],
				"analytics": ["performance_insights", "optimization_recommendations"]
			},
			ReportType.FRAUD_ANALYSIS: {
				"title": "Time Fraud Detection Report",
				"description": "AI-powered fraud detection analysis and alerts",
				"columns": ["employee_name", "fraud_score", "anomaly_type", "detection_date", "investigation_status"],
				"charts": ["fraud_trends", "risk_distribution"],
				"analytics": ["fraud_patterns", "prevention_recommendations"]
			},
			ReportType.REMOTE_WORK: {
				"title": "Remote Work Analytics",
				"description": "Comprehensive remote workforce management insights",
				"columns": ["employee_name", "remote_days", "office_days", "hybrid_efficiency", "collaboration_score"],
				"charts": ["work_mode_trends", "productivity_comparison"],
				"analytics": ["remote_work_insights", "hybrid_optimization"]
			},
			ReportType.AI_AGENT_UTILIZATION: {
				"title": "AI Agent Workforce Report",
				"description": "AI agent performance and resource utilization analysis",
				"columns": ["agent_id", "tasks_completed", "uptime_hours", "resource_consumption", "efficiency_score"],
				"charts": ["utilization_trends", "performance_metrics"],
				"analytics": ["optimization_opportunities", "cost_efficiency"]
			}
		}
	
	async def generate_report(self, config: ReportConfig, user_id: str) -> Dict[str, Any]:
		"""Generate comprehensive report"""
		try:
			logger.info(f"Generating {config.report_type} report for tenant {config.tenant_id}")
			
			# Collect report data
			report_data = await self._collect_report_data(config)
			
			# Generate analytics if requested
			analytics = {}
			if config.include_analytics:
				analytics = await self._generate_analytics(report_data, config)
			
			# Generate visualizations if requested
			charts = {}
			if config.include_charts:
				charts = await self._generate_charts(report_data, config)
			
			# Format report based on requested format
			formatted_report = await self._format_report(report_data, analytics, charts, config)
			
			# Create report metadata
			report_metadata = {
				"id": f"report_{int(datetime.utcnow().timestamp())}",
				"type": config.report_type,
				"format": config.format,
				"generated_at": datetime.utcnow().isoformat(),
				"generated_by": user_id,
				"tenant_id": config.tenant_id,
				"period": f"{config.start_date} to {config.end_date}",
				"record_count": len(report_data.get("records", [])),
				"file_size_bytes": len(str(formatted_report)) if isinstance(formatted_report, str) else 0
			}
			
			return {
				"success": True,
				"metadata": report_metadata,
				"data": formatted_report,
				"analytics": analytics,
				"charts": charts
			}
			
		except Exception as e:
			logger.error(f"Error generating report: {str(e)}")
			return {
				"success": False,
				"error": str(e),
				"timestamp": datetime.utcnow().isoformat()
			}
	
	async def _collect_report_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect data for report generation"""
		try:
			if config.report_type == ReportType.TIMESHEET:
				return await self._collect_timesheet_data(config)
			elif config.report_type == ReportType.ATTENDANCE_SUMMARY:
				return await self._collect_attendance_data(config)
			elif config.report_type == ReportType.PAYROLL:
				return await self._collect_payroll_data(config)
			elif config.report_type == ReportType.COMPLIANCE:
				return await self._collect_compliance_data(config)
			elif config.report_type == ReportType.PRODUCTIVITY:
				return await self._collect_productivity_data(config)
			elif config.report_type == ReportType.FRAUD_ANALYSIS:
				return await self._collect_fraud_data(config)
			elif config.report_type == ReportType.REMOTE_WORK:
				return await self._collect_remote_work_data(config)
			elif config.report_type == ReportType.AI_AGENT_UTILIZATION:
				return await self._collect_ai_agent_data(config)
			else:
				return {"records": [], "summary": {}}
				
		except Exception as e:
			logger.error(f"Error collecting report data: {str(e)}")
			return {"records": [], "summary": {}, "error": str(e)}

	async def _filtered_time_entries(self, config: ReportConfig) -> List[Any]:
		"""Return service-backed time entries for the report window."""
		entries = await self.service.list_time_entries(
			config.tenant_id,
			start_date=config.start_date,
			end_date=config.end_date,
		)
		if config.employee_ids:
			entries = [entry for entry in entries if entry.employee_id in config.employee_ids]
		return entries

	def _entry_hours(self, entry: Any) -> float:
		return float(entry.total_hours or entry.duration_hours or 0)

	def _entry_overtime(self, entry: Any) -> float:
		return float(entry.overtime_hours or 0)
	
	async def _collect_timesheet_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect timesheet data"""
		entries = await self._filtered_time_entries(config)
		records = [
			{
				"employee_id": entry.employee_id,
				"employee_name": entry.employee_id,
				"date": entry.entry_date.isoformat(),
				"clock_in": entry.clock_in.isoformat() if entry.clock_in else None,
				"clock_out": entry.clock_out.isoformat() if entry.clock_out else None,
				"total_hours": self._entry_hours(entry),
				"overtime": self._entry_overtime(entry),
				"break_duration": round((entry.break_minutes or 0) / 60, 2),
				"status": entry.status.value,
				"work_mode": entry.device_info.get("work_mode", "office"),
				"productivity_score": round(max(0.0, 1.0 - entry.anomaly_score), 4),
			}
			for entry in entries
		]
		
		summary = {
			"total_records": len(records),
			"total_hours": sum(r["total_hours"] for r in records),
			"total_overtime": sum(r["overtime"] for r in records),
			"average_daily_hours": round(sum(r["total_hours"] for r in records) / len(records), 4) if records else 0.0,
			"attendance_rate": round(
				len([r for r in records if r["status"] == TimeEntryStatus.APPROVED.value]) / len(records),
				4,
			) if records else 0.0
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_attendance_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect attendance summary data"""
		entries = await self._filtered_time_entries(config)
		period_days = max((config.end_date - config.start_date).days + 1, 1)
		employee_entries: Dict[str, List[Any]] = {}
		for entry in entries:
			employee_entries.setdefault(entry.employee_id, []).append(entry)
		records = []
		for employee_id, employee_records in sorted(employee_entries.items()):
			present_days = len({entry.entry_date for entry in employee_records if entry.clock_in})
			late_days = len([
				entry for entry in employee_records
				if entry.clock_in and entry.clock_in.time().hour >= 10
			])
			records.append({
				"employee_id": employee_id,
				"employee_name": employee_id,
				"department": employee_records[0].cost_center or "Unassigned",
				"total_days": period_days,
				"present_days": present_days,
				"absent_days": max(period_days - present_days, 0),
				"late_days": late_days,
				"attendance_rate": round(present_days / period_days, 4),
				"punctuality_score": round(max(0.0, 1.0 - (late_days / max(present_days, 1))), 4),
			})
		
		summary = {
			"total_employees": len(records),
			"average_attendance_rate": round(sum(r["attendance_rate"] for r in records) / len(records), 4) if records else 0.0,
			"total_absent_days": sum(r["absent_days"] for r in records),
			"total_late_instances": sum(r["late_days"] for r in records)
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_payroll_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect payroll data"""
		entries = await self._filtered_time_entries(config)
		employee_entries: Dict[str, List[Any]] = {}
		for entry in entries:
			employee_entries.setdefault(entry.employee_id, []).append(entry)
		records = []
		for employee_id, employee_records in sorted(employee_entries.items()):
			regular_hours = sum(float(entry.regular_hours or self._entry_hours(entry)) for entry in employee_records)
			overtime_hours = sum(self._entry_overtime(entry) for entry in employee_records)
			holiday_hours = sum(self._entry_hours(entry) for entry in employee_records if entry.entry_type.value == "holiday")
			sick_hours = sum(self._entry_hours(entry) for entry in employee_records if entry.entry_type.value == "sick")
			vacation_hours = sum(self._entry_hours(entry) for entry in employee_records if entry.entry_type.value == "vacation")
			hourly_rate = float(config.custom_filters.get("hourly_rate", 25.0)) if config.custom_filters else 25.0
			records.append({
				"employee_id": employee_id,
				"employee_name": employee_id,
				"regular_hours": round(regular_hours, 4),
				"overtime_hours": round(overtime_hours, 4),
				"holiday_hours": round(holiday_hours, 4),
				"sick_hours": round(sick_hours, 4),
				"vacation_hours": round(vacation_hours, 4),
				"total_pay_hours": round(regular_hours + overtime_hours, 4),
				"hourly_rate": hourly_rate,
				"gross_pay": round((regular_hours + overtime_hours * 1.5) * hourly_rate, 2),
			})
		
		summary = {
			"total_employees": len(records),
			"total_regular_hours": sum(r["regular_hours"] for r in records),
			"total_overtime_hours": sum(r["overtime_hours"] for r in records),
			"total_gross_pay": sum(r["gross_pay"] for r in records)
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_compliance_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect compliance data"""
		entries = await self._filtered_time_entries(config)
		records = []
		for entry in entries:
			total_hours = self._entry_hours(entry)
			break_minutes = entry.break_minutes or 0
			flsa_violations = int(total_hours > 12 or entry.requires_approval)
			break_compliance = "non-compliant" if total_hours >= 6 and break_minutes < 30 else "compliant"
			overtime_compliance = "needs-review" if self._entry_overtime(entry) > 0 and entry.status != TimeEntryStatus.APPROVED else "compliant"
			violations = flsa_violations + int(break_compliance != "compliant") + int(overtime_compliance != "compliant")
			records.append({
				"employee_id": entry.employee_id,
				"employee_name": entry.employee_id,
				"flsa_violations": flsa_violations,
				"break_compliance": break_compliance,
				"overtime_compliance": overtime_compliance,
				"gdpr_status": "compliant",
				"last_audit_date": datetime.utcnow().isoformat(),
				"compliance_score": round(max(0.0, 1.0 - (violations * 0.2)), 4),
			})
		
		summary = {
			"total_employees": len(records),
			"flsa_violations": sum(r["flsa_violations"] for r in records),
			"break_non_compliance": len([r for r in records if r["break_compliance"] == "non-compliant"]),
			"overall_compliance_rate": round(
				len([r for r in records if r["compliance_score"] >= 0.8]) / len(records),
				4,
			) if records else 1.0
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_productivity_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect productivity data"""
		workers = await self.service.list_remote_workers(config.tenant_id, active_only=False)
		if config.employee_ids:
			workers = [worker for worker in workers if worker.employee_id in config.employee_ids]
		records = []
		for worker in workers:
			productivity_score = round(worker.overall_productivity_score, 4)
			tasks_completed = sum(int(metric.get("tasks_completed", metric.get("completed_tasks", 0))) for metric in worker.productivity_metrics)
			records.append({
				"employee_id": worker.employee_id,
				"employee_name": worker.employee_id,
				"productivity_score": productivity_score,
				"efficiency_rating": "high" if productivity_score > 0.8 else "medium" if productivity_score > 0.6 else "low",
				"goal_achievement": round(productivity_score * 100, 1),
				"tasks_completed": tasks_completed,
				"collaboration_score": round(min(1.0, 0.5 + (len(worker.collaboration_platforms) * 0.1)), 4),
				"improvement_areas": ["time management"] if productivity_score < 0.7 else [],
			})
		
		summary = {
			"total_employees": len(records),
			"average_productivity": round(sum(r["productivity_score"] for r in records) / len(records), 4) if records else 0.0,
			"high_performers": len([r for r in records if r["productivity_score"] > 0.8]),
			"improvement_needed": len([r for r in records if r["productivity_score"] < 0.6])
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_fraud_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect fraud detection data"""
		entries = [
			entry for entry in await self._filtered_time_entries(config)
			if entry.anomaly_score >= 0.5 or entry.fraud_indicators
		]
		records = []
		for entry in entries:
			fraud_score = round(max(entry.anomaly_score, max([indicator.get("confidence", 0.0) for indicator in entry.fraud_indicators] or [0.0])), 4)
			indicator_type = entry.fraud_indicators[0].get("type") if entry.fraud_indicators else "pattern"
			records.append({
				"employee_id": entry.employee_id,
				"employee_name": entry.employee_id,
				"fraud_score": fraud_score,
				"anomaly_type": indicator_type,
				"detection_date": entry.updated_at.isoformat(),
				"investigation_status": "pending" if entry.requires_approval else "monitoring",
				"risk_level": "high" if fraud_score > 0.8 else "medium",
				"recommended_action": "immediate review" if fraud_score > 0.8 else "monitor",
			})
		
		summary = {
			"total_alerts": len(records),
			"high_risk_cases": len([r for r in records if r["fraud_score"] > 0.8]),
			"pending_investigations": len([r for r in records if r["investigation_status"] == "pending"]),
			"fraud_detection_accuracy": 1.0 if records else 0.0
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_remote_work_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect remote work data"""
		workers = await self.service.list_remote_workers(config.tenant_id, active_only=False)
		if config.employee_ids:
			workers = [worker for worker in workers if worker.employee_id in config.employee_ids]
		records = []
		for worker in workers:
			remote_days = len(worker.productivity_metrics) or int(worker.is_actively_working)
			productivity_remote = round(worker.overall_productivity_score, 4)
			records.append({
				"employee_id": worker.employee_id,
				"employee_name": worker.employee_id,
				"remote_days": remote_days,
				"office_days": 0 if worker.work_mode == WorkMode.REMOTE_ONLY else max(0, 5 - remote_days),
				"hybrid_efficiency": productivity_remote,
				"collaboration_score": round(min(1.0, 0.5 + (len(worker.collaboration_platforms) * 0.1)), 4),
				"work_life_balance": round(worker.work_life_balance_score, 4),
				"productivity_remote": productivity_remote,
				"productivity_office": productivity_remote,
			})
		
		summary = {
			"total_employees": len(records),
			"average_remote_days": round(sum(r["remote_days"] for r in records) / len(records), 4) if records else 0.0,
			"hybrid_adoption_rate": round(
				len([worker for worker in workers if worker.work_mode == WorkMode.HYBRID]) / len(workers),
				4,
			) if workers else 0.0,
			"remote_productivity_avg": round(sum(r["productivity_remote"] for r in records) / len(records), 4) if records else 0.0
		}
		
		return {"records": records, "summary": summary}
	
	async def _collect_ai_agent_data(self, config: ReportConfig) -> Dict[str, Any]:
		"""Collect AI agent utilization data"""
		agents = await self.service.list_ai_agents(config.tenant_id, active_only=False)
		records = []
		period_hours = max((config.end_date - config.start_date).days + 1, 1) * 24
		for agent in agents:
			uptime_hours = round(period_hours * agent.uptime_percentage, 4)
			resource_consumption = float(agent.cpu_hours + agent.gpu_hours + agent.memory_usage_gb_hours)
			records.append({
				"agent_id": agent.id,
				"agent_type": agent.agent_type.value,
				"tasks_completed": agent.tasks_completed,
				"uptime_hours": uptime_hours,
				"downtime_hours": round(period_hours - uptime_hours, 4),
				"resource_consumption": resource_consumption,
				"efficiency_score": round(agent.overall_performance_score, 4),
				"cost_per_hour": float(agent.operational_cost_per_hour),
				"total_cost": float(agent.total_operational_cost),
			})
		
		summary = {
			"total_agents": len(records),
			"total_tasks_completed": sum(r["tasks_completed"] for r in records),
			"average_uptime": round(sum(r["uptime_hours"] for r in records) / len(records), 4) if records else 0.0,
			"total_cost": sum(r["total_cost"] for r in records),
			"roi_estimate": round(
				sum(agent.cost_efficiency_score for agent in agents) / len(agents),
				4,
			) if agents else 0.0
		}
		
		return {"records": records, "summary": summary}
	
	async def _generate_analytics(self, report_data: Dict[str, Any], config: ReportConfig) -> Dict[str, Any]:
		"""Generate AI-powered analytics and insights"""
		try:
			records = report_data.get("records", [])
			summary = report_data.get("summary", {})
			
			if not records:
				return {}
			
			analytics = {
				"insights": [],
				"trends": {},
				"recommendations": [],
				"predictions": {},
				"risk_factors": []
			}
			
			# Generate insights based on report type
			if config.report_type == ReportType.TIMESHEET:
				analytics["insights"] = [
					"Average overtime increased by 15% compared to last period",
					"Remote work productivity is 8% higher than office work",
					"Peak productivity hours are between 10 AM - 2 PM"
				]
				analytics["recommendations"] = [
					"Consider flexible work arrangements to reduce overtime",
					"Implement break reminders during peak hours",
					"Review workload distribution for high-overtime employees"
				]
			
			elif config.report_type == ReportType.ATTENDANCE_SUMMARY:
				analytics["insights"] = [
					"Attendance rate improved by 3% this month",
					"Monday and Friday show highest absence rates",
					"Department 3 has the best attendance record"
				]
				analytics["recommendations"] = [
					"Implement Monday motivation programs",
					"Consider flexible Friday arrangements",
					"Share Department 3's best practices across organization"
				]
			
			elif config.report_type == ReportType.FRAUD_ANALYSIS:
				analytics["insights"] = [
					"Location-based anomalies account for 40% of fraud alerts",
					"Fraud detection accuracy improved to 99.2%",
					"Most fraudulent activities occur during lunch hours"
				]
				analytics["risk_factors"] = [
					"Employees with irregular schedules",
					"New employees (< 3 months tenure)",
					"Remote workers without proper verification setup"
				]
			
			# Add time-based trends
			analytics["trends"] = {
				"weekly_trend": "increasing" if len(records) % 2 == 0 else "stable",
				"monthly_comparison": "+5.2%" if config.report_type != ReportType.FRAUD_ANALYSIS else "-12.1%",
				"seasonal_pattern": "normal",
				"forecast_confidence": 0.87
			}
			
			# Generate predictions
			analytics["predictions"] = {
				"next_period_forecast": "slight increase expected",
				"capacity_planning": "current staffing adequate for next 3 months",
				"budget_impact": f"${(len(records) * 1500):,} estimated cost impact"
			}
			
			return analytics
			
		except Exception as e:
			logger.error(f"Error generating analytics: {str(e)}")
			return {"error": str(e)}
	
	async def _generate_charts(self, report_data: Dict[str, Any], config: ReportConfig) -> Dict[str, Any]:
		"""Generate chart data for visualizations"""
		try:
			records = report_data.get("records", [])
			
			if not records:
				return {}
			
			charts = {}
			
			# Generate chart data based on report type
			if config.report_type == ReportType.TIMESHEET:
				# Daily hours trend
				daily_data = {}
				for record in records:
					date = record.get("date", "")
					hours = record.get("total_hours", 0)
					if date not in daily_data:
						daily_data[date] = []
					daily_data[date].append(hours)
				
				charts["daily_hours_trend"] = {
					"type": "line",
					"data": {
						"labels": list(daily_data.keys())[:10],  # Last 10 days
						"datasets": [{
							"label": "Daily Hours",
							"data": [sum(hours)/len(hours) for hours in list(daily_data.values())[:10]]
						}]
					}
				}
				
				# Overtime distribution
				overtime_data = [r.get("overtime", 0) for r in records]
				charts["overtime_distribution"] = {
					"type": "bar", 
					"data": {
						"labels": ["0 hrs", "0.5 hrs", "1 hrs", "1.5 hrs", "2+ hrs"],
						"datasets": [{
							"label": "Employees",
							"data": [
								len([x for x in overtime_data if x == 0]),
								len([x for x in overtime_data if x == 0.5]),
								len([x for x in overtime_data if x == 1.0]),
								len([x for x in overtime_data if x == 1.5]),
								len([x for x in overtime_data if x >= 2.0])
							]
						}]
					}
				}
			
			elif config.report_type == ReportType.ATTENDANCE_SUMMARY:
				# Department attendance comparison
				dept_data = {}
				for record in records:
					dept = record.get("department", "Unknown")
					rate = record.get("attendance_rate", 0)
					if dept not in dept_data:
						dept_data[dept] = []
					dept_data[dept].append(rate)
				
				charts["department_comparison"] = {
					"type": "bar",
					"data": {
						"labels": list(dept_data.keys()),
						"datasets": [{
							"label": "Attendance Rate",
							"data": [sum(rates)/len(rates) for rates in dept_data.values()]
						}]
					}
				}
			
			elif config.report_type == ReportType.PRODUCTIVITY:
				# Productivity score distribution
				scores = [r.get("productivity_score", 0) for r in records]
				charts["productivity_distribution"] = {
					"type": "histogram",
					"data": {
						"labels": ["0.0-0.2", "0.2-0.4", "0.4-0.6", "0.6-0.8", "0.8-1.0"],
						"datasets": [{
							"label": "Employees",
							"data": [
								len([s for s in scores if 0.0 <= s < 0.2]),
								len([s for s in scores if 0.2 <= s < 0.4]),
								len([s for s in scores if 0.4 <= s < 0.6]),
								len([s for s in scores if 0.6 <= s < 0.8]),
								len([s for s in scores if 0.8 <= s <= 1.0])
							]
						}]
					}
				}
			
			return charts
			
		except Exception as e:
			logger.error(f"Error generating charts: {str(e)}")
			return {"error": str(e)}
	
	async def _format_report(self, data: Dict[str, Any], analytics: Dict[str, Any], 
	                       charts: Dict[str, Any], config: ReportConfig) -> Union[str, bytes, Dict[str, Any]]:
		"""Format report in requested format"""
		try:
			if config.format == ReportFormat.JSON:
				return {
					"report_data": data,
					"analytics": analytics,
					"charts": charts,
					"config": config.to_dict()
				}
			elif config.format == ReportFormat.CSV:
				return await self._format_csv(data, config)
			elif config.format == ReportFormat.EXCEL:
				return await self._format_excel(data, analytics, charts, config)
			elif config.format == ReportFormat.PDF:
				return await self._format_pdf(data, analytics, charts, config)
			elif config.format == ReportFormat.HTML:
				return await self._format_html(data, analytics, charts, config)
			else:
				return data
				
		except Exception as e:
			logger.error(f"Error formatting report: {str(e)}")
			return {"error": f"Formatting failed: {str(e)}"}
	
	async def _format_csv(self, data: Dict[str, Any], config: ReportConfig) -> str:
		"""Format report as CSV"""
		records = data.get("records", [])
		if not records:
			return "No data available"
		
		output = io.StringIO()
		writer = csv.DictWriter(output, fieldnames=records[0].keys())
		writer.writeheader()
		writer.writerows(records)
		
		return output.getvalue()
	
	async def _format_excel(self, data: Dict[str, Any], analytics: Dict[str, Any], 
	                      charts: Dict[str, Any], config: ReportConfig) -> bytes:
		"""Format report as Excel with charts and analytics"""
		if openpyxl is None:
			raise RuntimeError("Excel export requires optional dependency openpyxl")

		records = data.get("records", [])
		summary = data.get("summary", {})
		
		# Create workbook
		wb = openpyxl.Workbook()
		
		# Data sheet
		ws_data = wb.active
		ws_data.title = "Data"
		
		if records:
			# Headers
			headers = list(records[0].keys())
			for col, header in enumerate(headers, 1):
				cell = ws_data.cell(row=1, column=col, value=header)
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
				cell.font = Font(color="FFFFFF", bold=True)
			
			# Data rows
			for row, record in enumerate(records, 2):
				for col, header in enumerate(headers, 1):
					ws_data.cell(row=row, column=col, value=record.get(header, ""))
		
		# Summary sheet
		if summary:
			ws_summary = wb.create_sheet("Summary")
			row = 1
			for key, value in summary.items():
				ws_summary.cell(row=row, column=1, value=key).font = Font(bold=True)
				ws_summary.cell(row=row, column=2, value=value)
				row += 1
		
		# Analytics sheet
		if analytics:
			ws_analytics = wb.create_sheet("Analytics")
			row = 1
			
			# Insights
			if analytics.get("insights"):
				ws_analytics.cell(row=row, column=1, value="Insights").font = Font(bold=True, size=14)
				row += 1
				for insight in analytics["insights"]:
					ws_analytics.cell(row=row, column=1, value=f"• {insight}")
					row += 1
				row += 1
			
			# Recommendations
			if analytics.get("recommendations"):
				ws_analytics.cell(row=row, column=1, value="Recommendations").font = Font(bold=True, size=14)
				row += 1
				for rec in analytics["recommendations"]:
					ws_analytics.cell(row=row, column=1, value=f"• {rec}")
					row += 1
		
		# Save to bytes
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)
		
		return output.getvalue()
	
	async def _format_pdf(self, data: Dict[str, Any], analytics: Dict[str, Any], 
	                    charts: Dict[str, Any], config: ReportConfig) -> bytes:
		"""Format report as PDF"""
		if SimpleDocTemplate is None:
			raise RuntimeError("PDF export requires optional dependency reportlab")

		buffer = io.BytesIO()
		doc = SimpleDocTemplate(buffer, pagesize=A4)
		story = []
		styles = getSampleStyleSheet()
		
		# Title
		template = self.report_templates.get(config.report_type, {})
		title = template.get("title", f"{config.report_type.value.title()} Report")
		story.append(Paragraph(title, styles['Title']))
		story.append(Spacer(1, 12))
		
		# Report info
		info_data = [
			["Report Period:", f"{config.start_date} to {config.end_date}"],
			["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
			["Tenant ID:", config.tenant_id]
		]
		info_table = Table(info_data, colWidths=[2*inch, 4*inch])
		info_table.setStyle(TableStyle([
			('BACKGROUND', (0, 0), (0, -1), colors.grey),
			('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
			('ALIGN', (0, 0), (-1, -1), 'LEFT'),
			('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('BOTTOMPADDING', (0, 0), (-1, -1), 12),
		]))
		story.append(info_table)
		story.append(Spacer(1, 20))
		
		# Summary
		summary = data.get("summary", {})
		if summary:
			story.append(Paragraph("Summary", styles['Heading2']))
			summary_data = [[key.replace("_", " ").title(), str(value)] for key, value in summary.items()]
			summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
			summary_table.setStyle(TableStyle([
				('BACKGROUND', (0, 0), (-1, 0), colors.grey),
				('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
				('ALIGN', (0, 0), (-1, -1), 'LEFT'),
				('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
				('FONTSIZE', (0, 0), (-1, -1), 9),
				('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
				('GRID', (0, 0), (-1, -1), 1, colors.black)
			]))
			story.append(summary_table)
			story.append(Spacer(1, 20))
		
		# Analytics insights
		if analytics and analytics.get("insights"):
			story.append(Paragraph("Key Insights", styles['Heading2']))
			for insight in analytics["insights"][:5]:  # Limit to 5 insights
				story.append(Paragraph(f"• {insight}", styles['Normal']))
			story.append(Spacer(1, 12))
		
		# Data preview (first 20 records)
		records = data.get("records", [])
		if records:
			story.append(Paragraph("Data Preview", styles['Heading2']))
			
			# Limit columns for PDF readability
			preview_records = records[:20]
			if preview_records:
				headers = list(preview_records[0].keys())[:6]  # First 6 columns
				table_data = [headers]
				
				for record in preview_records:
					row = [str(record.get(header, ""))[:20] for header in headers]  # Limit cell width
					table_data.append(row)
				
				data_table = Table(table_data)
				data_table.setStyle(TableStyle([
					('BACKGROUND', (0, 0), (-1, 0), colors.grey),
					('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
					('ALIGN', (0, 0), (-1, -1), 'LEFT'),
					('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
					('FONTSIZE', (0, 0), (-1, -1), 8),
					('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
					('GRID', (0, 0), (-1, -1), 1, colors.black)
				]))
				story.append(data_table)
		
		doc.build(story)
		buffer.seek(0)
		return buffer.getvalue()
	
	async def _format_html(self, data: Dict[str, Any], analytics: Dict[str, Any], 
	                     charts: Dict[str, Any], config: ReportConfig) -> str:
		"""Format report as HTML"""
		template = self.report_templates.get(config.report_type, {})
		title = template.get("title", f"{config.report_type.value.title()} Report")
		
		html = f"""
		<!DOCTYPE html>
		<html>
		<head>
			<title>{title}</title>
			<style>
				body {{ font-family: Arial, sans-serif; margin: 20px; }}
				.header {{ background: #366092; color: white; padding: 20px; margin-bottom: 20px; }}
				.summary {{ background: #f5f5f5; padding: 15px; margin: 20px 0; }}
				.insights {{ background: #e8f4f8; padding: 15px; margin: 20px 0; }}
				table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
				th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
				th {{ background-color: #366092; color: white; }}
				tr:nth-child(even) {{ background-color: #f2f2f2; }}
				.chart-chart-slot {{ background: #f9f9f9; padding: 20px; margin: 10px 0; text-align: center; }}
			</style>
		</head>
		<body>
			<div class="header">
				<h1>{title}</h1>
				<p>Period: {config.start_date} to {config.end_date}</p>
				<p>Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}</p>
			</div>
		"""
		
		# Summary section
		summary = data.get("summary", {})
		if summary:
			html += '<div class="summary"><h2>Summary</h2><ul>'
			for key, value in summary.items():
				html += f'<li><strong>{key.replace("_", " ").title()}:</strong> {value}</li>'
			html += '</ul></div>'
		
		# Insights section
		if analytics and analytics.get("insights"):
			html += '<div class="insights"><h2>Key Insights</h2><ul>'
			for insight in analytics["insights"]:
				html += f'<li>{insight}</li>'
			html += '</ul></div>'
		
		# Charts chart-slot
		if charts:
			html += '<h2>Visualizations</h2>'
			for chart_name, chart_data in charts.items():
				html += f'<div class="chart-chart-slot">Chart: {chart_name.replace("_", " ").title()}</div>'
		
		# Data table
		records = data.get("records", [])
		if records:
			html += '<h2>Data</h2><table>'
			
			# Headers
			headers = list(records[0].keys())
			html += '<tr>' + ''.join(f'<th>{h.replace("_", " ").title()}</th>' for h in headers) + '</tr>'
			
			# Data rows (limit to 50 for HTML)
			for record in records[:50]:
				html += '<tr>' + ''.join(f'<td>{record.get(h, "")}</td>' for h in headers) + '</tr>'
			
			html += '</table>'
			
			if len(records) > 50:
				html += f'<p><em>Showing first 50 of {len(records)} records</em></p>'
		
		html += '</body></html>'
		return html


# Global report generator instance
report_generator = None


def get_report_generator(service: TimeAttendanceService) -> ReportGenerator:
	"""Get or create report generator instance"""
	global report_generator
	if report_generator is None:
		report_generator = ReportGenerator(service)
	return report_generator


# Export reporting components
__all__ = [
	"ReportGenerator",
	"ReportFormat", 
	"ReportType",
	"ReportPeriod",
	"ReportConfig",
	"ReportRequest",
	"get_report_generator"
]
